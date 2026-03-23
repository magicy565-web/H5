"""
Storage module – persists leads as JSON and generates Excel daily reports.
"""
from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from monitor.config import DB_DIR, OUTPUT_DIR

if TYPE_CHECKING:
    pass  # future: from monitor.analyzer import Lead

logger = logging.getLogger(__name__)

# ── Ensure directories exist ──────────────────────────────────────────
DB_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ── JSON helpers ──────────────────────────────────────────────────────

def _get_leads_file():
    """Get the current LEADS_FILE (may change per industry)."""
    from monitor.config import LEADS_FILE
    return LEADS_FILE


def load_leads() -> list[dict]:
    """Load existing leads from the active industry's leads file."""
    leads_file = _get_leads_file()
    if not leads_file.exists():
        logger.debug("Leads file %s does not exist yet.", leads_file)
        return []
    try:
        data = json.loads(leads_file.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return data
        logger.warning("leads file root is not a list – returning empty.")
        return []
    except (json.JSONDecodeError, OSError) as exc:
        logger.warning("Failed to read leads file: %s", exc)
        return []


def save_leads(leads: list[dict]) -> None:
    """Write *leads* list to the active industry's leads file."""
    leads_file = _get_leads_file()
    leads_file.write_text(
        json.dumps(leads, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info("Saved %d leads to %s.", len(leads), leads_file)


def _lead_to_dict(lead: Any) -> dict:
    """Convert a Lead dataclass (or plain dict) to a dict suitable for
    JSON storage.  Accepts both dicts and objects with a ``to_dict`` or
    ``__dict__`` interface."""
    if isinstance(lead, dict):
        return lead
    # dataclass / object with to_dict()
    if hasattr(lead, "to_dict"):
        return lead.to_dict()
    if hasattr(lead, "__dict__"):
        return {k: v for k, v in lead.__dict__.items() if not k.startswith("_")}
    return dict(lead)


def append_leads(new_leads: list) -> list[dict]:
    """Load existing leads, merge *new_leads* (avoiding duplicates by
    ``contentHash``), save back, and return the full list.

    *new_leads* may contain Lead dataclass instances or plain dicts.
    Returns the merged list of dicts.
    """
    existing = load_leads()
    existing_hashes: set[str] = set()
    for entry in existing:
        h = entry.get("contentHash") or entry.get("content_hash") or ""
        if h:
            existing_hashes.add(h)

    added = 0
    for lead in new_leads:
        d = _lead_to_dict(lead)
        h = d.get("contentHash") or d.get("content_hash") or ""
        if not h:
            # No hash available - compute one from url+title as fallback
            url = d.get("sourceUrl") or d.get("source_url") or d.get("url") or ""
            title = d.get("title") or ""
            if url or title:
                import hashlib
                h = hashlib.md5(f"{url}|{title}".lower().strip().encode()).hexdigest()[:12]
                d["contentHash"] = h
        if h and h in existing_hashes:
            continue
        existing.append(d)
        if h:
            existing_hashes.add(h)
        added += 1

    logger.info("Appended %d new leads (skipped %d duplicates).",
                added, len(new_leads) - added)
    save_leads(existing)
    return existing


# ── Excel report ──────────────────────────────────────────────────────

_HEADERS = [
    "ID",
    "来源",
    "发现时间",
    "标题",
    "意向评分",
    "买家国家",
    "买家名称",
    "设备规格",
    "紧迫度",
    "中文摘要",
    "推荐动作",
    "联系方式",
    "原文链接",
]

_FIELD_MAP: list[str] = [
    "id",
    "source",
    "discoveredAt",
    "title",
    "intentScore",
    "buyerCountry",
    "buyerName",
    "machineSpecs",
    "urgency",
    "summaryZh",
    "recommendedAction",
    "contactInfo",
    "sourceUrl",
]

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_excel(leads: list[dict], date_str: str, industry: str = "") -> Path:
    """Generate an Excel report and return the output file path.

    The report is saved to ``OUTPUT_DIR/采购意向日报_{industry}_{date_str}.xlsx``.
    """
    # Sort by intentScore descending (treat missing as 0)
    sorted_leads = sorted(
        leads,
        key=lambda d: d.get("intentScore", d.get("intent_score", 0)) or 0,
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "采购意向"

    # ── Header row ────────────────────────────────────────────────
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, lead in enumerate(sorted_leads, start=2):
        for col_idx, field in enumerate(_FIELD_MAP, start=1):
            # Support both camelCase and snake_case keys
            value = lead.get(field)
            if value is None:
                snake = _camel_to_snake(field)
                value = lead.get(snake, "")
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

    # ── Auto-width columns ────────────────────────────────────────
    for col_idx in range(1, len(_HEADERS) + 1):
        max_len = 0
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=1, max_row=ws.max_row):
            for cell in row:
                try:
                    cell_len = len(str(cell.value or ""))
                    # CJK characters are roughly double-width
                    cjk_count = sum(1 for c in str(cell.value or "") if ord(c) > 0x4E00)
                    cell_len += cjk_count
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ── Save ──────────────────────────────────────────────────────
    tag = f"_{industry}" if industry else ""
    time_str = datetime.now().strftime("%H%M")
    filename = f"采购意向日报{tag}_{date_str}_{time_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    try:
        wb.save(str(filepath))
    except PermissionError:
        logger.error(
            "Cannot save Excel report to %s – file may be locked by another process. "
            "Trying alternative filename …", filepath,
        )
        alt_filename = f"采购意向日报{tag}_{date_str}_{time_str}_alt.xlsx"
        filepath = OUTPUT_DIR / alt_filename
        try:
            wb.save(str(filepath))
        except PermissionError:
            logger.error("Still cannot save Excel report to %s – giving up.", filepath)
            return filepath
    except OSError as exc:
        logger.error("Failed to save Excel report to %s: %s", filepath, exc)
        return filepath
    logger.info("Excel report saved to %s (%d rows).", filepath, len(sorted_leads))
    return filepath


# ── Utilities ─────────────────────────────────────────────────────────

def _camel_to_snake(name: str) -> str:
    """Convert a camelCase name to snake_case."""
    import re
    s = re.sub(r"([A-Z])", r"_\1", name)
    return s.lower().lstrip("_")
